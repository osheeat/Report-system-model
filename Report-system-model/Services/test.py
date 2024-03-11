import tkinter as tk
from tkinter import filedialog
import openpyxl
import pandas as pd
import psycopg2
from psycopg2 import sql
from sqlalchemy import create_engine

unique_values = []
conn = psycopg2.connect(
        host="localhost",
        database="rsm",
        user="postgres",
        password="postgres"
    )

file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
excel_file = file_path
df = pd.read_excel(excel_file)

# Выбор необходимого столбца
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active
row_number = 1
max_col = sheet.max_column

table_name = ''


def create_array():
    import pandas as pd

    # Замените 'your_excel_file.xlsx' на имя вашего файла Excel
    excel_file = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])

    # Загрузите данные из Excel в DataFrame
    df = pd.read_excel(excel_file)

    # Преобразуйте DataFrame в массив массивов (список списков)
    data_as_list_of_lists = df.values.tolist()
    return data_as_list_of_lists


def keyfigure_filiing():
    create_array()
    import psycopg2
    from psycopg2 import sql

    # Замените параметрами подключения на ваши
    conn_params = {
        'host': 'localhost',
        'database': 'rsm9',
        'user': 'postgres',
        'password': 'postgres',
        'port': '5432'
    }

    # Замените 'your_table' на имя вашей таблицы в базе данных
    table_name = 'keyfigure'

    data_as_list_of_lists = create_array()
    # Создайте соединение с базой данных
    conn = psycopg2.connect(**conn_params)

    # Создайте курсор для выполнения SQL-запросов
    cursor = conn.cursor()

    # Подготовьте SQL-запрос для вставки данных в таблицу
    columns = ','.join(['sap_code', 'full_name', 'short_name', 'value_type_id', 'currency_unit_id', 'method_of_obtaining_id',
                        'keyfigure_category_id', 'load_time_id', 'upload_deadline_id', 'report_usage_indicator_id',
                        'data_status_id', 'formation_methodology',
                        'indicator_generation_method_id'])  # Замените на реальные названия столбцов
    placeholders = ','.join(['%s' for _ in range(len(data_as_list_of_lists[0]))])
    insert_query = sql.SQL(
        f"INSERT INTO {table_name} (sap_code, full_name, short_name, value_type_id, currency_unit_id, method_of_obtaining_id, keyfigure_category_id, load_time_id, upload_deadline_id, report_usage_indicator_id, data_status_id, formation_methodology, indicator_generation_method_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)").format(
        sql.Identifier(table_name),
        sql.SQL(columns),
        sql.SQL(placeholders)
    )

    # Вставьте данные в таблицу
    cursor.executemany(insert_query, data_as_list_of_lists)

    # Сделайте коммит, чтобы изменения вступили в силу
    conn.commit()

    # Закройте курсор и соединение
    cursor.close()
    conn.close()

def fill_single_tables():
    try:
        # Проходимся по всем столбцам
        for col_num in range(1, max_col + 1):
            # Получаем буквенное представление номера столбца (A, B, C, ..., AA, AB, ...)
            col_letter = openpyxl.utils.get_column_letter(col_num)

            # Снова неизменяемая часть не считая названия столбца в sql-запросе
            cell_value = sheet[f'{col_letter}{row_number}'].value
            print(f'Текст из ячейки {col_letter}{row_number}: {cell_value}')

            # Выберите столбец, в котором нужно найти уникальные строки
            column = cell_value
            unique_values = df[column].drop_duplicates().tolist()

            print(unique_values)

            # Получаем значение из ячейки (например, из ячейки A1)
            cell = f'{col_letter}1'
            cell_value = sheet[cell].value

            # Создаем строку с текстом из ячейки
            table_name = str(cell_value)
            # Создание курсора
            cur = conn.cursor()
            # Список значений, которые вы хотите добавить
            values = unique_values

            # Название столбца, который вы хотите проверить
            column_name = "value"
            cursor = conn.cursor()
            query = f"SELECT EXISTS (SELECT 1 FROM information_schema.columns WHERE table_name = '{table_name}' AND column_name = '{column_name}')"
            cursor.execute(query)
            column_exists = cursor.fetchone()[0]

            if column_exists:
                sql = f"INSERT INTO {table_name} (value) VALUES (%s)"
                # Использование цикла для добавления каждого значения в базу данных
                for значение in values:
                    cur.execute(sql, (значение,))
                conn.commit()

            else:
                sql = f"INSERT INTO {table_name} (title) VALUES (%s)"
                # Использование цикла для добавления каждого значения в базу данных
                for значение in values:
                    cur.execute(sql, (значение,))
                conn.commit()

            print(table_name, 'заполнилась успешно')

    except:
        print(table_name, 'не заполнилась')

    finally:
        print('Конец выполнения')

def indicator_source_system_filling():
    conn = psycopg2.connect(
        dbname="rsm9",
        user="postgres",
        password="postgres",
        host="localhost"
    )

    cur = conn.cursor()

    df = pd.read_excel("C:/Users/oleg9/study/Диплом/Service/indicator_source_system.xlsx")

    for index, row in df.iterrows():
        # Получите название из Excel
        keyfigure = row['keyfigure_id']
        cur.execute("SELECT id FROM keyfigure WHERE sap_code = %s", (keyfigure,))
        result = cur.fetchone()

        company = row['company_id']
        cur.execute("SELECT id FROM company WHERE title = %s", (company,))
        result2 = cur.fetchone()

        source_system = row['source_system_id']
        cur.execute("SELECT id FROM source_system WHERE title = %s", (source_system,))
        result3 = cur.fetchone()

        if result and result2 and result3:
            # Если идентификатор найден, используйте его для вставки данных
            id_названия = result[0]
            id_company = result2[0]
            id_ss = result3[0]
            # Вставьте данные в вашу таблицу, заменив 'ваша_таблица' и 'столбцы' на соответствующие значения
            cur.execute(
                "INSERT INTO indicator_source_system (keyfigure_id, company_id, source_system_id) VALUES (%s, %s, %s)",
                (id_названия, id_company, id_ss))

    # Не забудьте зафиксировать изменения и закрыть соединение
    conn.commit()
    cur.close()
    conn.close()

def release_indicator_source_system_filling():
    conn = psycopg2.connect(
        dbname="rsm9",
        user="postgres",
        password="postgres",
        host="localhost"
    )

    cur = conn.cursor()

    df = pd.read_excel("C:/Users/oleg9/study/Диплом/Service/release_indicator_source_system.xlsx")

    for index, row in df.iterrows():
        # Получите название из Excel
        keyfigure = row['keyfigure_id']
        cur.execute("SELECT id FROM keyfigure WHERE sap_code = %s", (keyfigure,))
        result = cur.fetchone()

        company = row['company_id']
        cur.execute("SELECT id FROM company WHERE title = %s", (company,))
        result2 = cur.fetchone()

        source_system = row['source_system_id']
        cur.execute("SELECT id FROM source_system WHERE title = %s", (source_system,))
        result3 = cur.fetchone()


        if result and result2 and result3:
            # Если идентификатор найден, используйте его для вставки данных
            id_названия = result[0]
            id_company = result2[0]
            id_ss = result3[0]
            # Вставьте данные в вашу таблицу, заменив 'ваша_таблица' и 'столбцы' на соответствующие значения
            cur.execute(
                "INSERT INTO release_indicator_source_system (keyfigure_id, company_id, source_system_id, release_id) VALUES (%s, %s, %s, %s)",
                (id_названия, id_company, id_ss, row['release_id']))

    # Не забудьте зафиксировать изменения и закрыть соединение
    conn.commit()
    cur.close()
    conn.close()

import psycopg2
from psycopg2 import sql

def report_filling():
    # Замените параметрами подключения на ваши
    conn_params = {
        'host': 'localhost',
        'database': 'rsm9',
        'user': 'postgres',
        'password': 'postgres',
        'port': '5432'
    }

    # Замените 'your_table' на имя вашей таблицы в базе данных
    table_name = 'report'

    # Создайте соединение с базой данных
    conn = psycopg2.connect(**conn_params)

    # Создайте курсор для выполнения SQL-запросов
    cursor = conn.cursor()

    for index, row in df.iterrows():
        keyfigure = row['keyfigure_id']
        cursor.execute("SELECT id FROM keyfigure WHERE sap_code = %s", (keyfigure,))
        result = cursor.fetchone()

        if result:
            keyfig = result[0]
            # Подготовьте SQL-запрос для вставки данных в таблицу
            insert_query = sql.SQL(
                f"INSERT INTO {table_name} (report_indicator_id, report_code_id,report_id_id, report_title_id, business_process_id, formation_frequency_id,analytical_feature_id, data_status_id, keyfigure_category_id, release_id, keyfigure_id) "
                f"VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
            ).format(
                sql.Identifier(table_name)
            )

            # Вставьте данные в таблицу
            cursor.execute(insert_query, (
                row['report_indicator_id'], row['report_code_id'], row['report_id_id'], row['report_title_id'],
                row['business_process_id'], row['formation_frequency_id'], row['analytical_feature_id'],
                row['data_status_id'], row['keyfigure_category_id'], row['release_id'], keyfig
            ))

    # Сделайте коммит, чтобы изменения вступили в силу
    conn.commit()

    # Закройте курсор и соединение
    cursor.close()
    conn.close()



fill_single_tables()
# keyfigure_filiing()
# report_filling()
# indicator_source_system_filling()
# release_indicator_source_system_filling()