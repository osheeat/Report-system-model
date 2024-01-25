import pandas as pd
import psycopg2
import openpyxl
import pandas as pd
import psycopg2
from sqlalchemy import create_engine

unique_values = []
conn = psycopg2.connect(
        host="localhost",
        database="tryRSM5",
        user="postgres",
        password="postgres"
    )


def company_filling():
    excel_file = 'C:/Users/oleg9/study/Диплом/1.xlsx'
    df = pd.read_excel(excel_file)

    # Выбор необходимого столбца
    workbook = openpyxl.load_workbook('C:/Users/oleg9/study/Диплом/1.xlsx')
    sheet = workbook.active
    row_number = 1
    column_letter = 'A'

    cell_value = sheet[f'{column_letter}{row_number}'].value
    print(f'Текст из ячейки {column_letter}{row_number}: {cell_value}')

    # Выберите столбец, в котором нужно найти уникальные строки
    column = cell_value
    unique_values = df[column].drop_duplicates().tolist()

    print(unique_values)

    table_name = 'company'
    # Создание курсора
    cur = conn.cursor()
    # Список значений, которые вы хотите добавить
    values = unique_values
    sql = f"INSERT INTO {table_name} (title) VALUES (%s)"
    # Использование цикла для добавления каждого значения в базу данных
    for значение in values:
        cur.execute(sql, (значение,))

    # Подтверждение изменений и закрытие соединения
    conn.commit()


def currency_unit_filling():
    excel_file = 'C:/Users/oleg9/study/Диплом/1.xlsx'
    df = pd.read_excel(excel_file)

    # Выбор необходимого столбца
    workbook = openpyxl.load_workbook('C:/Users/oleg9/study/Диплом/1.xlsx')
    sheet = workbook.active
    row_number = 1
    column_letter = 'B'

    cell_value = sheet[f'{column_letter}{row_number}'].value
    print(f'Текст из ячейки {column_letter}{row_number}: {cell_value}')

    # Выберите столбец, в котором нужно найти уникальные строки
    column = cell_value
    unique_values = df[column].drop_duplicates().tolist()

    print(unique_values)

    table_name = cell_value
    # Создание курсора
    cur = conn.cursor()
    # Список значений, которые вы хотите добавить
    values = unique_values
    sql = f"INSERT INTO {table_name} (value) VALUES (%s)"
    # Использование цикла для добавления каждого значения в базу данных
    for значение in values:
        cur.execute(sql, (значение,))

    # Подтверждение изменений и закрытие соединения
    conn.commit()

def data_status_filling():
    excel_file = 'C:/Users/oleg9/study/Диплом/1.xlsx'
    df = pd.read_excel(excel_file)

    # Выбор необходимого столбца
    workbook = openpyxl.load_workbook('C:/Users/oleg9/study/Диплом/1.xlsx')
    sheet = workbook.active
    row_number = 1
    column_letter = 'C'

    cell_value = sheet[f'{column_letter}{row_number}'].value
    print(f'Текст из ячейки {column_letter}{row_number}: {cell_value}')

    # Выберите столбец, в котором нужно найти уникальные строки
    column = cell_value
    unique_values = df[column].drop_duplicates().tolist()

    print(unique_values)

    table_name = cell_value
    # Создание курсора
    cur = conn.cursor()
    # Список значений, которые вы хотите добавить
    values = unique_values
    sql = f"INSERT INTO {table_name} (value) VALUES (%s)"
    # Использование цикла для добавления каждого значения в базу данных
    for значение in values:
        cur.execute(sql, (значение,))

    # Подтверждение изменений и закрытие соединения
    conn.commit()


def indicator_generation_method_filling():
    excel_file = 'C:/Users/oleg9/study/Диплом/1.xlsx'
    df = pd.read_excel(excel_file)

    # Выбор необходимого столбца
    workbook = openpyxl.load_workbook('C:/Users/oleg9/study/Диплом/1.xlsx')
    sheet = workbook.active
    row_number = 1
    column_letter = 'D'

    cell_value = sheet[f'{column_letter}{row_number}'].value
    print(f'Текст из ячейки {column_letter}{row_number}: {cell_value}')

    # Выберите столбец, в котором нужно найти уникальные строки
    column = cell_value
    unique_values = df[column].drop_duplicates().tolist()

    print(unique_values)

    table_name = cell_value
    # Создание курсора
    cur = conn.cursor()
    # Список значений, которые вы хотите добавить
    values = unique_values
    sql = f"INSERT INTO {table_name} (value) VALUES (%s)"
    # Использование цикла для добавления каждого значения в базу данных
    for значение in values:
        cur.execute(sql, (значение,))

    # Подтверждение изменений и закрытие соединения
    conn.commit()


def keyfigure_category_filling():
    excel_file = 'C:/Users/oleg9/study/Диплом/1.xlsx'
    df = pd.read_excel(excel_file)

    # Выбор необходимого столбца
    workbook = openpyxl.load_workbook('C:/Users/oleg9/study/Диплом/1.xlsx')
    sheet = workbook.active
    row_number = 1
    column_letter = 'F'

    cell_value = sheet[f'{column_letter}{row_number}'].value
    print(f'Текст из ячейки {column_letter}{row_number}: {cell_value}')

    # Выберите столбец, в котором нужно найти уникальные строки
    column = cell_value
    unique_values = df[column].drop_duplicates().tolist()

    print(unique_values)

    table_name = cell_value
    # Создание курсора
    cur = conn.cursor()
    # Список значений, которые вы хотите добавить
    values = unique_values
    sql = f"INSERT INTO {table_name} (value) VALUES (%s)"
    # Использование цикла для добавления каждого значения в базу данных
    for значение in values:
        cur.execute(sql, (значение,))

    # Подтверждение изменений и закрытие соединения
    conn.commit()


def load_time_filling():
    excel_file = 'C:/Users/oleg9/study/Диплом/1.xlsx'
    df = pd.read_excel(excel_file)

    # Выбор необходимого столбца
    workbook = openpyxl.load_workbook('C:/Users/oleg9/study/Диплом/1.xlsx')
    sheet = workbook.active
    row_number = 1
    column_letter = 'G'

    cell_value = sheet[f'{column_letter}{row_number}'].value
    print(f'Текст из ячейки {column_letter}{row_number}: {cell_value}')

    # Выберите столбец, в котором нужно найти уникальные строки
    column = cell_value
    unique_values = df[column].drop_duplicates().tolist()

    print(unique_values)

    table_name = cell_value
    # Создание курсора
    cur = conn.cursor()
    # Список значений, которые вы хотите добавить
    values = unique_values
    sql = f"INSERT INTO {table_name} (value) VALUES (%s)"
    # Использование цикла для добавления каждого значения в базу данных
    for значение in values:
        cur.execute(sql, (значение,))

    # Подтверждение изменений и закрытие соединения
    conn.commit()


def method_of_obtaining_filling():
    excel_file = 'C:/Users/oleg9/study/Диплом/1.xlsx'
    df = pd.read_excel(excel_file)

    # Выбор необходимого столбца
    workbook = openpyxl.load_workbook('C:/Users/oleg9/study/Диплом/1.xlsx')
    sheet = workbook.active
    row_number = 1
    column_letter = 'H'

    cell_value = sheet[f'{column_letter}{row_number}'].value
    print(f'Текст из ячейки {column_letter}{row_number}: {cell_value}')

    # Выберите столбец, в котором нужно найти уникальные строки
    column = cell_value
    unique_values = df[column].drop_duplicates().tolist()

    print(unique_values)

    table_name = cell_value
    # Создание курсора
    cur = conn.cursor()
    # Список значений, которые вы хотите добавить
    values = unique_values
    sql = f"INSERT INTO {table_name} (value) VALUES (%s)"
    # Использование цикла для добавления каждого значения в базу данных
    for значение in values:
        cur.execute(sql, (значение,))

    # Подтверждение изменений и закрытие соединения
    conn.commit()


def release_filling():
    excel_file = 'C:/Users/oleg9/study/Диплом/1.xlsx'
    df = pd.read_excel(excel_file)

    # Выбор необходимого столбца
    workbook = openpyxl.load_workbook('C:/Users/oleg9/study/Диплом/1.xlsx')
    sheet = workbook.active
    row_number = 1
    column_letter = 'I'

    cell_value = sheet[f'{column_letter}{row_number}'].value
    print(f'Текст из ячейки {column_letter}{row_number}: {cell_value}')

    # Выберите столбец, в котором нужно найти уникальные строки
    column = cell_value
    unique_values = df[column].drop_duplicates().tolist()

    print(unique_values)

    table_name = cell_value
    # Создание курсора
    cur = conn.cursor()
    # Список значений, которые вы хотите добавить
    values = unique_values
    sql = f"INSERT INTO {table_name} (value) VALUES (%s)"
    # Использование цикла для добавления каждого значения в базу данных
    for значение in values:
        cur.execute(sql, (значение,))

    # Подтверждение изменений и закрытие соединения
    conn.commit()


def report_usage_indicator_filling():
    excel_file = 'C:/Users/oleg9/study/Диплом/1.xlsx'
    df = pd.read_excel(excel_file)

    # Выбор необходимого столбца
    workbook = openpyxl.load_workbook('C:/Users/oleg9/study/Диплом/1.xlsx')
    sheet = workbook.active
    row_number = 1
    column_letter = 'J'

    cell_value = sheet[f'{column_letter}{row_number}'].value
    print(f'Текст из ячейки {column_letter}{row_number}: {cell_value}')

    # Выберите столбец, в котором нужно найти уникальные строки
    column = cell_value
    unique_values = df[column].drop_duplicates().tolist()

    print(unique_values)

    table_name = cell_value
    # Создание курсора
    cur = conn.cursor()
    # Список значений, которые вы хотите добавить
    values = unique_values
    sql = f"INSERT INTO {table_name} (value) VALUES (%s)"
    # Использование цикла для добавления каждого значения в базу данных
    for значение in values:
        cur.execute(sql, (значение,))

    # Подтверждение изменений и закрытие соединения
    conn.commit()



def source_system_filling():
    excel_file = 'C:/Users/oleg9/study/Диплом/1.xlsx'
    df = pd.read_excel(excel_file)

    # Выбор необходимого столбца
    workbook = openpyxl.load_workbook('C:/Users/oleg9/study/Диплом/1.xlsx')
    sheet = workbook.active
    row_number = 1
    column_letter = 'K'

    cell_value = sheet[f'{column_letter}{row_number}'].value
    print(f'Текст из ячейки {column_letter}{row_number}: {cell_value}')

    # Выберите столбец, в котором нужно найти уникальные строки
    column = cell_value
    unique_values = df[column].drop_duplicates().tolist()

    print(unique_values)

    table_name = cell_value
    # Создание курсора
    cur = conn.cursor()
    # Список значений, которые вы хотите добавить
    values = unique_values
    sql = f"INSERT INTO {table_name} (title) VALUES (%s)"
    # Использование цикла для добавления каждого значения в базу данных
    for значение in values:
        cur.execute(sql, (значение,))

    # Подтверждение изменений и закрытие соединения
    conn.commit()



def upload_deadline_filling():
    excel_file = 'C:/Users/oleg9/study/Диплом/1.xlsx'
    df = pd.read_excel(excel_file)

    # Выбор необходимого столбца
    workbook = openpyxl.load_workbook('C:/Users/oleg9/study/Диплом/1.xlsx')
    sheet = workbook.active
    row_number = 1
    column_letter = 'L'

    cell_value = sheet[f'{column_letter}{row_number}'].value
    print(f'Текст из ячейки {column_letter}{row_number}: {cell_value}')

    # Выберите столбец, в котором нужно найти уникальные строки
    column = cell_value
    unique_values = df[column].drop_duplicates().tolist()

    print(unique_values)

    table_name = cell_value
    # Создание курсора
    cur = conn.cursor()
    # Список значений, которые вы хотите добавить
    values = unique_values
    sql = f"INSERT INTO {table_name} (value) VALUES (%s)"
    # Использование цикла для добавления каждого значения в базу данных
    for значение in values:
        cur.execute(sql, (значение,))

    # Подтверждение изменений и закрытие соединения
    conn.commit()



def value_type_filling():
    excel_file = 'C:/Users/oleg9/study/Диплом/1.xlsx'
    df = pd.read_excel(excel_file)

    # Выбор необходимого столбца
    workbook = openpyxl.load_workbook('C:/Users/oleg9/study/Диплом/1.xlsx')
    sheet = workbook.active
    row_number = 1
    column_letter = 'M'

    cell_value = sheet[f'{column_letter}{row_number}'].value
    print(f'Текст из ячейки {column_letter}{row_number}: {cell_value}')

    # Выберите столбец, в котором нужно найти уникальные строки
    column = cell_value
    unique_values = df[column].drop_duplicates().tolist()

    print(unique_values)

    table_name = cell_value
    # Создание курсора
    cur = conn.cursor()
    # Список значений, которые вы хотите добавить
    values = unique_values
    sql = f"INSERT INTO {table_name} (value) VALUES (%s)"
    # Использование цикла для добавления каждого значения в базу данных
    for значение in values:
        cur.execute(sql, (значение,))

    # Подтверждение изменений и закрытие соединения
    conn.commit()

def create_array():
    import pandas as pd

    # Замените 'your_excel_file.xlsx' на имя вашего файла Excel
    excel_file = 'C:/Users/oleg9/study/Диплом/2.xlsx'

    # Загрузите данные из Excel в DataFrame
    df = pd.read_excel(excel_file)

    # Преобразуйте DataFrame в массив массивов (список списков)
    data_as_list_of_lists = df.values.tolist()

    print(data_as_list_of_lists)
    return data_as_list_of_lists


def keyfigure_filiing():
    create_array()
    import psycopg2
    from psycopg2 import sql

    # Замените параметрами подключения на ваши
    conn_params = {
        'host': 'localhost',
        'database': 'tryRSM5',
        'user': 'postgres',
        'password': 'postgres',
        'port': '5432'
    }

    # Замените 'your_table' на имя вашей таблицы в базе данных
    table_name = 'keyfigure'

    data_as_list_of_lists = create_array()
    # Создайте соединение с базой данных
    conn = psycopg2.connect(**conn_params)

    # Создайте курсор для выполнения SQL-запросов
    cursor = conn.cursor()

    # Подготовьте SQL-запрос для вставки данных в таблицу
    columns = ','.join(['sap_code', 'full_name', 'short_name', 'value_type_id', 'currency_unit_id', 'method_of_obtaining_id',
                        'keyfigure_category_id', 'load_time_id', 'upload_deadline_id', 'report_usage_indicator_id',
                        'data_status_id', 'formation_methodology',
                        'indicator_generation_method_id'])  # Замените на реальные названия столбцов
    placeholders = ','.join(['%s' for _ in range(len(data_as_list_of_lists[0]))])
    insert_query = sql.SQL(
        f"INSERT INTO {table_name} (sap_code, full_name, short_name, value_type_id, currency_unit_id, method_of_obtaining_id, keyfigure_category_id, load_time_id, upload_deadline_id, report_usage_indicator_id, data_status_id, formation_methodology, indicator_generation_method_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)").format(
        sql.Identifier(table_name),
        sql.SQL(columns),
        sql.SQL(placeholders)
    )

    # Вставьте данные в таблицу
    cursor.executemany(insert_query, data_as_list_of_lists)

    # Сделайте коммит, чтобы изменения вступили в силу
    conn.commit()

    # Закройте курсор и соединение
    cursor.close()
    conn.close()


def filling_db():
    # company_filling()
    # currency_unit_filling()
    # data_status_filling()
    # indicator_generation_method_filling()
    # keyfigure_category_filling()
    # load_time_filling()
    # method_of_obtaining_filling()
    # release_filling()
    # report_usage_indicator_filling()
    # source_system_filling()
    # upload_deadline_filling()
    # value_type_filling()
    keyfigure_filiing()



filling_db()




